"""
Merge a scraper.py run's output folder into the main output/ dataset, and
rebuild progress_tracker.xlsx to reflect the new totals.

Usage:
    python merge_and_track.py output_maharashtra
    python merge_and_track.py output_maharashtra --force   # overwrite any
                                                             # districts already
                                                             # present in output/

By default, if a district in the new run is already present in output/
(e.g. you accidentally re-ran a completed district), that district's new
rows are skipped and a warning is printed, to avoid silently duplicating
data. Pass --force to replace the old rows for that district instead.

After merging, the source folder (e.g. output_maharashtra/) is deleted
unless --keep-temp is passed.
"""

import argparse
import csv
import json
import shutil
import sys
import urllib.request
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

REPO_ROOT = Path(__file__).parent
OUTPUT_DIR = REPO_ROOT / "output"
ALL_DISTRICTS_CACHE = REPO_ROOT / "recon" / "all_states_districts.json"
TRACKER_PATH = REPO_ROOT / "progress_tracker.xlsx"

STATE_IDS = {
    "ANDAMAN AND NICOBAR ISLANDS": "1", "ANDHRA PRADESH": "2", "ARUNACHAL PRADESH": "3",
    "ASSAM": "4", "BIHAR": "5", "CHANDIGARH": "6", "CHHATTISGARH": "33",
    "DADRA AND NAGAR HAVELI": "7", "DAMAN AND DIU": "8", "DELHI": "9", "GOA": "10",
    "GUJARAT": "11", "HARYANA": "12", "HIMACHAL PRADESH": "13", "JAMMU AND KASHMIR": "14",
    "JHARKHAND": "35", "KARNATAKA": "15", "KERALA": "16", "LADAKH": "37",
    "LAKSHADWEEP": "17", "MADHYA PRADESH": "18", "MAHARASHTRA": "19", "MANIPUR": "20",
    "MEGHALAYA": "21", "MIZORAM": "22", "NAGALAND": "23", "ODISHA": "24", "OTHER": "99",
    "PONDICHERRY": "25", "PUNJAB": "26", "RAJASTHAN": "27", "SIKKIM": "28",
    "TAMILNADU": "29", "TELANGANA": "36", "TRIPURA": "30", "UTTAR PRADESH": "31",
    "UTTARAKHAND": "34", "WEST BENGAL": "32",
}


def fetch_all_states_districts():
    """(Re)build the master state->district reference list from the open,
    unauthenticated loadDistByStateId API. No CAPTCHA involved."""
    print("Fetching full state/district reference list from the open API...")
    all_rows = []
    for sname, sid in STATE_IDS.items():
        url = f"https://psaonline.utiitsl.com/PanPSACenters/forms/loadDistByStateId/{sid}"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        for d in data:
            raw = d["distName"]
            all_rows.append({"state_id": sid, "state": sname, "district_raw": raw,
                              "district": raw.strip()})
    ALL_DISTRICTS_CACHE.parent.mkdir(parents=True, exist_ok=True)
    with ALL_DISTRICTS_CACHE.open("w", encoding="utf-8") as f:
        json.dump(all_rows, f, indent=2, ensure_ascii=False)
    return all_rows


def load_all_states_districts():
    if ALL_DISTRICTS_CACHE.exists():
        with ALL_DISTRICTS_CACHE.open(encoding="utf-8") as f:
            return json.load(f)
    return fetch_all_states_districts()


def write_dataset(records, out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)
    json_path = out_dir / "pan_centers.json"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)

    all_keys = []
    for r in records:
        for k in r.keys():
            if k not in all_keys:
                all_keys.append(k)
    csv_path = out_dir / "pan_centers.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=all_keys)
        writer.writeheader()
        writer.writerows(records)

    import pandas as pd
    pd.DataFrame(records).to_excel(out_dir / "pan_centers.xlsx", index=False)


def rebuild_tracker(records):
    counts = Counter((r["_search_state"], r["_search_district"]) for r in records)
    all_districts = load_all_states_districts()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Districts"
    ws.append(["State", "District", "Completed", "Record Count"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")

    green = PatternFill("solid", fgColor="C6EFCE")
    for d in all_districts:
        key = (d["state"], d["district"])
        n = counts.get(key)
        done = n is not None
        ws.append([d["state"], d["district"], "☑" if done else "☐", n if done else ""])
        if done:
            for cell in ws[ws.max_row]:
                cell.fill = green

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col, width in zip("ABCD", [28, 24, 12, 14]):
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet("States Summary")
    ws2.append(["State", "Total Districts", "Completed Districts", "Total Records", "Fully Complete"])
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9D9D9")

    states_order = []
    seen = set()
    for d in all_districts:
        if d["state"] not in seen:
            seen.add(d["state"])
            states_order.append(d["state"])

    for state in states_order:
        dists = [d for d in all_districts if d["state"] == state]
        total = len(dists)
        completed = sum(1 for d in dists if (state, d["district"]) in counts)
        total_records = sum(counts.get((state, d["district"]), 0) for d in dists)
        fully = "☑" if completed == total else "☐"
        ws2.append([state, total, completed, total_records, fully])
        if completed == total:
            for cell in ws2[ws2.max_row]:
                cell.fill = green
        elif completed > 0:
            for cell in ws2[ws2.max_row]:
                cell.fill = PatternFill("solid", fgColor="FFEB9C")

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    for col, width in zip("ABCDE", [28, 16, 20, 16, 16]):
        ws2.column_dimensions[col].width = width

    wb.save(TRACKER_PATH)
    print(f"Updated {TRACKER_PATH} "
          f"({sum(1 for d in all_districts if (d['state'], d['district']) in counts)}/"
          f"{len(all_districts)} district entries completed).")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("new_output", help="Path to the new run's --out-dir folder (e.g. output_maharashtra)")
    parser.add_argument("--force", action="store_true",
                         help="Overwrite existing data for districts that already exist in output/")
    parser.add_argument("--keep-temp", action="store_true",
                         help="Don't delete the source folder after merging")
    args = parser.parse_args()

    new_dir = Path(args.new_output)
    new_json = new_dir / "pan_centers.json"
    if not new_json.exists():
        print(f"No pan_centers.json found in {new_dir} — nothing to merge.")
        sys.exit(1)

    with new_json.open(encoding="utf-8") as f:
        new_records = json.load(f)

    if OUTPUT_DIR.joinpath("pan_centers.json").exists():
        with OUTPUT_DIR.joinpath("pan_centers.json").open(encoding="utf-8") as f:
            existing = json.load(f)
    else:
        existing = []

    existing_keys = {(r["_search_state"], r["_search_district"]) for r in existing}
    new_keys = {(r["_search_state"], r["_search_district"]) for r in new_records}
    overlap = existing_keys & new_keys

    if overlap:
        if args.force:
            print(f"--force set: replacing {len(overlap)} district(s) already in output/: "
                  f"{sorted(overlap)}")
            existing = [r for r in existing if (r["_search_state"], r["_search_district"]) not in overlap]
        else:
            print(f"WARNING: {len(overlap)} district(s) already exist in output/ — skipping "
                  f"their new rows to avoid duplicates (use --force to overwrite): {sorted(overlap)}")
            new_records = [r for r in new_records if (r["_search_state"], r["_search_district"]) not in overlap]

    merged = existing + new_records
    write_dataset(merged, OUTPUT_DIR)
    print(f"output/ now has {len(merged)} total records "
          f"(+{len(new_records)} new from {new_dir}).")

    rebuild_tracker(merged)

    if not args.keep_temp:
        shutil.rmtree(new_dir, ignore_errors=True)
        print(f"Removed temporary folder {new_dir}.")


if __name__ == "__main__":
    main()
